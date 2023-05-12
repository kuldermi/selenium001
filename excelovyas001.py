import openpyxl
import sqlite3

conn = sqlite3.connect("example.db")

cursor = conn.cursor()

# cursor.execute('''CREATE TABLE users (id INTEGER PRIMARY KEY, name TEXT, email TEXT )''')
# cursor.execute('''INSERT INTO users (name, email) VALUES ('Khokan', 'abc.gmail.com')''')
# conn.commit()
cursor.execute("SELECT * FROM users")

rows = cursor.fetchall()
for row in rows:
    print(row)



#
# workbook = openpyxl.load_workbook("Wave2API.xlsx")
#
# worksheet = workbook["Squad 1"]
# worksheet.cell(row= 10, column = 7).value = "Testing hoche"
# value = worksheet.cell(row= 10, column= 7).value
# workbook.save("Wave2API.xlsx")
#
# print(value)
#

